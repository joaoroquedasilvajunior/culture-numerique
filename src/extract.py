"""
Fonctions d'extraction des tableaux ISQ vers des structures Python normalisées.

Conventions
-----------
Chaque extracteur prend un chemin de fichier .xlsx et retourne un dict.
Les valeurs textuelles spéciales de l'ISQ sont systématiquement traduites :
    '..'   -> None  (donnée non disponible)
    '...'  -> None  (sans objet — typiquement séries terminées)
    '--'   -> 0.0   (donnée infime, pour les TCM)
    '-'    -> 0.0   (néant ou zéro)
    'F'    -> None  (donnée peu fiable, ne peut être diffusée)

Les libellés contiennent souvent des espaces insécables \\xa0 (hiérarchie SCIAN)
et des préfixes en points (ex. '..Albums physiques'). On normalise en supprimant
ces marqueurs après en avoir extrait le niveau hiérarchique.
"""

from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook


# ---------- Helpers ----------

def _clean_label(s):
    if s is None:
        return ''
    s = str(s).replace('\xa0', '').replace('\n', ' ').strip()
    s = re.sub(r'^\.+\s*', '', s)  # remove leading sub-level markers
    return s


def _to_num(x):
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if s in ('', '..', '...', 'F'):
            return None
        if s in ('--', '-'):
            return 0.0
        try:
            return float(s.replace(',', '.'))
        except (ValueError, TypeError):
            return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _hierarchy_level(s):
    """Return indentation level inferred from leading non-breaking spaces or dots."""
    if s is None:
        return 0
    raw = str(s)
    # 6 \xa0 per level (ISQ convention)
    nbsp = len(raw) - len(raw.lstrip('\xa0'))
    if nbsp > 0:
        return nbsp // 6
    # ".." per level (some files use this convention instead)
    dots = len(raw) - len(raw.lstrip('.'))
    return dots // 2


def _is_terminated(s):
    return '(Terminé)' in str(s or '')


# ---------- Extractors ----------

def extract_part_qc(path: Path) -> dict:
    """
    Tableau « Part des interprètes du Québec dans la consommation d'enregistrements musicaux »
    Layout : col A = libellé, col B = % cette semaine, col D = % cumulatif YTD.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    cibles = {
        'streaming': "Écoutes sur les services de diffusion de musique en continu",
        'albums_total': 'Ensemble des albums',
        'albums_physiques': 'Albums sur support physique',
        'albums_numeriques': 'Albums en fichier numérique',
        'pistes_numeriques': 'Pistes en fichier numérique',
    }
    indicateurs = {}
    for row in ws.iter_rows(min_row=8, max_row=20, values_only=True):
        label = _clean_label(row[0])
        for k, target in cibles.items():
            if label.startswith(target):
                indicateurs[k] = {
                    'libelle': target,
                    'semaine_pct': _to_num(row[1]),
                    'cumul_ytd_pct': _to_num(row[3]),
                }
    return {'periode': str(periode).strip(), 'indicateurs': indicateurs}


def extract_volume_musique(path: Path) -> dict:
    """
    Tableau « Consommation d'enregistrements musicaux selon le type de produit »
    Layout : A=libellé, B=cette sem (k), D=var sem-1 (%), F=var an-1 sem (%),
             H=cumul YTD (k), J=var cumul an-1 (%).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    cibles = {
        'streaming': "Écoutes sur les services de diffusion de musique en continu",
        'albums_total': 'Ensemble des albums',
        'albums_physiques': 'Albums sur support physique',
        'cd': 'Disques compacts',
        'vinyle': 'Disques vinyle',
        'albums_numeriques': 'Albums en fichier numérique',
        'pistes_numeriques': 'Pistes en fichier numérique',
    }
    indicateurs = {}
    for row in ws.iter_rows(min_row=8, max_row=22, values_only=True):
        label = _clean_label(row[0])
        for k, target in cibles.items():
            if label.startswith(target):
                indicateurs[k] = {
                    'libelle': target,
                    'semaine': _to_num(row[1]),
                    'var_sem_prec_pct': _to_num(row[3]),
                    'var_an_prec_sem_pct': _to_num(row[5]),
                    'cumul_ytd': _to_num(row[7]),
                    'var_cumul_an_prec_pct': _to_num(row[9]),
                }
    return {'periode': str(periode).strip(), 'indicateurs': indicateurs}


def extract_cinema_pays(path: Path) -> dict:
    """
    Tableau « Résultats d'exploitation cinéma selon pays d'origine »
    Layout : A=pays, B=assistance sem (n), D=% sem, F=var sem-1 %, H=var an-1 sem %,
             J=cumul YTD (n), L=% cumul, N=var cumul an-1 %.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    pays_cibles = ['États-Unis', 'France', 'Grande-Bretagne', 'Québec',
                   'Canada', 'Autres pays', 'Total']
    pays_data = []
    for row in ws.iter_rows(min_row=10, max_row=22, values_only=True):
        label = _clean_label(row[0])
        if label in pays_cibles:
            pays_data.append({
                'pays': label,
                'assistance_semaine': _to_num(row[1]),
                'pct_semaine': _to_num(row[3]),
                'var_sem_prec_pct': _to_num(row[5]),
                'var_an_prec_sem_pct': _to_num(row[7]),
                'assistance_cumul_ytd': _to_num(row[9]),
                'pct_cumul_ytd': _to_num(row[11]),
                'var_cumul_an_prec_pct': _to_num(row[13]),
            })
    return {'periode': str(periode).strip(), 'pays': pays_data}


def extract_palmares(path: Path) -> list[dict]:
    """
    Tableau « Palmarès des enregistrements musicaux »
    Layout : A=rang (entier), B=interprète, C=provenance.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    top = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            rang = int(row[0])
        except (TypeError, ValueError):
            continue
        top.append({
            'rang': rang,
            'interprete': str(row[1]).strip() if row[1] else '',
            'provenance': str(row[2]).strip() if row[2] else '',
        })
    return top


def extract_evolution(path: Path) -> dict:
    """
    Tableau « Évolution de statistiques clés de la culture et des communications »
    Layout : A=domaine (header rows), B=indicateur, C=unité, puis pairs (year, blank).
    Header année en row 4 sous la forme '2020(ou 2020-2021)'.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Find year columns from row 4
    annees = []
    year_cols = []
    for c in ws[4]:
        v = c.value
        if v and 'ou' in str(v):
            try:
                yr = int(str(v).split('(')[0].strip())
                annees.append(yr)
                year_cols.append(c.column)
            except (ValueError, IndexError):
                pass

    # Whitelist of indicators to extract (clé -> début du libellé)
    cibles = {
        "scene_representations": "Représentations payantes",
        "scene_assistance": "Assistance totale des représentations payantes",
        "scene_revenus": "Revenus de billetterie excluant les taxes",
        "cine_projections": "Projections dans les cinémas et ciné-parcs",
        "cine_assistance": "Assistance dans les cinémas et ciné-parcs",
        "cine_recettes": "Recettes de billetterie dans les cinémas",
        "musique_streaming": "Nombre d'écoutes sur les services de diffusion",
        "musique_ventes": "Nombre d'enregistrements vendus",
        "livre_ventes": "Ventes finales de livres neufs",
        "biblio_usagers": "Usagers inscrits dans les bibliothèques",
        "musees_freq": "Fréquentation totale dans les institutions",
        "depenses_pub": "Dépenses en culture de l'administration",
        "depenses_menages": "Ensemble des dépenses pour la culture",
    }

    indicateurs = {}
    for row in ws.iter_rows(min_row=6, max_row=40, values_only=True):
        label_b = row[1] if len(row) > 1 else None
        if not label_b:
            continue
        label = str(label_b).replace('\xa0', '').strip()
        for key, target in cibles.items():
            if label.startswith(target):
                serie = []
                for yr, col in zip(annees, year_cols):
                    val = _to_num(row[col - 1]) if (col - 1) < len(row) else None
                    serie.append({'annee': yr, 'valeur': val})
                indicateurs[key] = {
                    'libelle': label[:80],
                    'unite': str(row[2]).strip() if row[2] else '',
                    'serie': serie,
                }
                break
    return {'annees': annees, 'indicateurs': indicateurs}


def extract_emplois_eerh(path: Path) -> list[dict]:
    """
    Tableau ISQ 2576 « Emplois salariés EERH ».
    Pour chaque industrie, on calcule la variation annuelle (Jan -> Déc) et le TCM cumulé.
    Retourne une liste de dicts : un par industrie active.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    MOIS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
            'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    month_cols = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25]  # 1-indexed

    rows_iter = list(ws.iter_rows(values_only=True))
    out = []
    i = 7  # data rows start at row 8 (0-indexed = 7)
    while i < len(rows_iter):
        r = rows_iter[i]
        col0, col1 = r[0], r[1]
        if (col0 and isinstance(col0, str) and col0.strip()
                and col1 == 'n'):
            industry_raw = col0
            niveau = _hierarchy_level(industry_raw)
            terminee = _is_terminated(industry_raw)

            scian_match = re.search(r'\(([\d,\s]+)\)\s*(?:\(Terminé\))?\s*$',
                                    industry_raw.replace('\xa0', '').strip())
            scian = scian_match.group(1).replace(' ', '') if scian_match else ''
            label_clean = re.sub(r'\s*\([\d,\s]+\)\s*(?:\(Terminé\))?\s*$', '',
                                 industry_raw.replace('\xa0', '').strip()).strip()

            # n values
            n_series = [_to_num(r[c - 1]) for c in month_cols]
            # TCM row (next)
            tcm_row = rows_iter[i + 1] if i + 1 < len(rows_iter) else (None,) * 26
            tcm_series = [_to_num(tcm_row[c - 1]) for c in month_cols]

            n_jan = n_series[0]
            n_dec = n_series[-1]
            if n_jan is not None and n_dec is not None and n_jan != 0:
                variation_pct = (n_dec - n_jan) / n_jan * 100
            else:
                variation_pct = None

            out.append({
                'scian': scian,
                'industrie': label_clean,
                'niveau': niveau,
                'serie_terminee': terminee,
                'mois': MOIS,
                'n_serie': n_series,
                'tcm_serie': tcm_series,
                'n_janvier': n_jan,
                'n_decembre': n_dec,
                'variation_pct': variation_pct,
            })
            i += 2
        else:
            i += 1

    return out


# ---------- Registry ----------

EXTRACTORS = {
    'extract_part_qc': extract_part_qc,
    'extract_volume_musique': extract_volume_musique,
    'extract_cinema_pays': extract_cinema_pays,
    'extract_palmares': extract_palmares,
    'extract_evolution': extract_evolution,
    'extract_emplois_eerh': extract_emplois_eerh,
}
