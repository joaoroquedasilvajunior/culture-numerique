"""
Fonctions d'extraction des tableaux ISQ vers des structures Python normalisées.

Conventions
-----------
Chaque extracteur prend un chemin de fichier .xlsx et retourne un dict.
Les valeurs textuelles spéciales de l'ISQ sont systématiquement traduites :
    '..'   -> None  (donnée non disponible)
    '...'  -> None  (sans objet — typiquement séries terminées)
    '--'   -> 0.0   (donnée infime, pour les TCM)
    '-'    -> 0.0   (néant ou zéro)
    'F'    -> None  (donnée peu fiable, ne peut être diffusée)

Les libellés contiennent souvent des espaces insécables \\xa0 (hiérarchie SCIAN)
et des préfixes en points (ex. '..Albums physiques'). On normalise en supprimant
ces marqueurs après en avoir extrait le niveau hiérarchique.
"""

from __future__ import annotations
import csv
import io
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook


# ---------- Helpers ----------

def _clean_label(s):
    if s is None:
        return ''
    s = str(s).replace('\xa0', '').replace('\n', ' ').strip()
    s = re.sub(r'^\.+\s*', '', s)  # remove leading sub-level markers
    return s


def _to_num(x):
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if s in ('', '..', '...', 'F'):
            return None
        if s in ('--', '-'):
            return 0.0
        try:
            return float(s.replace(',', '.'))
        except (ValueError, TypeError):
            return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _hierarchy_level(s):
    """Return indentation level inferred from leading non-breaking spaces or dots."""
    if s is None:
        return 0
    raw = str(s)
    # 6 \xa0 per level (ISQ convention)
    nbsp = len(raw) - len(raw.lstrip('\xa0'))
    if nbsp > 0:
        return nbsp // 6
    # ".." per level (some files use this convention instead)
    dots = len(raw) - len(raw.lstrip('.'))
    return dots // 2


def _is_terminated(s):
    return '(Terminé)' in str(s or '')


# ---------- Extractors ----------

def extract_part_qc(path: Path) -> dict:
    """
    Tableau « Part des interprètes du Québec dans la consommation d'enregistrements musicaux »
    Layout : col A = libellé, col B = % cette semaine, col D = % cumulatif YTD.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    cibles = {
        'streaming': "Écoutes sur les services de diffusion de musique en continu",
        'albums_total': 'Ensemble des albums',
        'albums_physiques': 'Albums sur support physique',
        'albums_numeriques': 'Albums en fichier numérique',
        'pistes_numeriques': 'Pistes en fichier numérique',
    }
    indicateurs = {}
    for row in ws.iter_rows(min_row=8, max_row=20, values_only=True):
        label = _clean_label(row[0])
        for k, target in cibles.items():
            if label.startswith(target):
                indicateurs[k] = {
                    'libelle': target,
                    'semaine_pct': _to_num(row[1]),
                    'cumul_ytd_pct': _to_num(row[3]),
                }
    return {'periode': str(periode).strip(), 'indicateurs': indicateurs}


def extract_volume_musique(path: Path) -> dict:
    """
    Tableau « Consommation d'enregistrements musicaux selon le type de produit »
    Layout : A=libellé, B=cette sem (k), D=var sem-1 (%), F=var an-1 sem (%),
             H=cumul YTD (k), J=var cumul an-1 (%).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    cibles = {
        'streaming': "Écoutes sur les services de diffusion de musique en continu",
        'albums_total': 'Ensemble des albums',
        'albums_physiques': 'Albums sur support physique',
        'cd': 'Disques compacts',
        'vinyle': 'Disques vinyle',
        'albums_numeriques': 'Albums en fichier numérique',
        'pistes_numeriques': 'Pistes en fichier numérique',
    }
    indicateurs = {}
    for row in ws.iter_rows(min_row=8, max_row=22, values_only=True):
        label = _clean_label(row[0])
        for k, target in cibles.items():
            if label.startswith(target):
                indicateurs[k] = {
                    'libelle': target,
                    'semaine': _to_num(row[1]),
                    'var_sem_prec_pct': _to_num(row[3]),
                    'var_an_prec_sem_pct': _to_num(row[5]),
                    'cumul_ytd': _to_num(row[7]),
                    'var_cumul_an_prec_pct': _to_num(row[9]),
                }
    return {'periode': str(periode).strip(), 'indicateurs': indicateurs}


def extract_cinema_pays(path: Path) -> dict:
    """
    Tableau « Résultats d'exploitation cinéma selon pays d'origine »
    Layout : A=pays, B=assistance sem (n), D=% sem, F=var sem-1 %, H=var an-1 sem %,
             J=cumul YTD (n), L=% cumul, N=var cumul an-1 %.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    periode = ws.cell(row=4, column=1).value or ''

    pays_cibles = ['États-Unis', 'France', 'Grande-Bretagne', 'Québec',
                   'Canada', 'Autres pays', 'Total']
    pays_data = []
    for row in ws.iter_rows(min_row=10, max_row=22, values_only=True):
        label = _clean_label(row[0])
        if label in pays_cibles:
            pays_data.append({
                'pays': label,
                'assistance_semaine': _to_num(row[1]),
                'pct_semaine': _to_num(row[3]),
                'var_sem_prec_pct': _to_num(row[5]),
                'var_an_prec_sem_pct': _to_num(row[7]),
                'assistance_cumul_ytd': _to_num(row[9]),
                'pct_cumul_ytd': _to_num(row[11]),
                'var_cumul_an_prec_pct': _to_num(row[13]),
            })
    return {'periode': str(periode).strip(), 'pays': pays_data}


def extract_palmares(path: Path) -> list[dict]:
    """
    Tableau « Palmarès des enregistrements musicaux »
    Layout : A=rang (entier), B=interprète, C=provenance.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    top = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            rang = int(row[0])
        except (TypeError, ValueError):
            continue
        top.append({
            'rang': rang,
            'interprete': str(row[1]).strip() if row[1] else '',
            'provenance': str(row[2]).strip() if row[2] else '',
        })
    return top


def extract_evolution(path: Path) -> dict:
    """
    Tableau « Évolution de statistiques clés de la culture et des communications »
    Layout : A=domaine (header rows), B=indicateur, C=unité, puis pairs (year, blank).
    Header année en row 4 sous la forme '2020(ou 2020-2021)'.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Find year columns from row 4
    annees = []
    year_cols = []
    for c in ws[4]:
        v = c.value
        if v and 'ou' in str(v):
            try:
                yr = int(str(v).split('(')[0].strip())
                annees.append(yr)
                year_cols.append(c.column)
            except (ValueError, IndexError):
                pass

    # Whitelist of indicators to extract (clé -> début du libellé)
    cibles = {
        "scene_representations": "Représentations payantes",
        "scene_assistance": "Assistance totale des représentations payantes",
        "scene_revenus": "Revenus de billetterie excluant les taxes",
        "cine_projections": "Projections dans les cinémas et ciné-parcs",
        "cine_assistance": "Assistance dans les cinémas et ciné-parcs",
        "cine_recettes": "Recettes de billetterie dans les cinémas",
        "musique_streaming": "Nombre d'écoutes sur les services de diffusion",
        "musique_ventes": "Nombre d'enregistrements vendus",
        "livre_ventes": "Ventes finales de livres neufs",
        "biblio_usagers": "Usagers inscrits dans les bibliothèques",
        "musees_freq": "Fréquentation totale dans les institutions",
        "depenses_pub": "Dépenses en culture de l'administration",
        "depenses_menages": "Ensemble des dépenses pour la culture",
    }

    indicateurs = {}
    for row in ws.iter_rows(min_row=6, max_row=40, values_only=True):
        label_b = row[1] if len(row) > 1 else None
        if not label_b:
            continue
        label = str(label_b).replace('\xa0', '').strip()
        for key, target in cibles.items():
            if label.startswith(target):
                serie = []
                for yr, col in zip(annees, year_cols):
                    val = _to_num(row[col - 1]) if (col - 1) < len(row) else None
                    serie.append({'annee': yr, 'valeur': val})
                indicateurs[key] = {
                    'libelle': label[:80],
                    'unite': str(row[2]).strip() if row[2] else '',
                    'serie': serie,
                }
                break
    return {'annees': annees, 'indicateurs': indicateurs}


def extract_emplois_eerh(path: Path) -> list[dict]:
    """
    Tableau ISQ 2576 « Emplois salariés EERH » — fichier mensuel MM3.

    Tolérant à l'année partielle : quand l'ISQ bascule sur l'année en cours,
    les mois non encore publiés portent la marque « (À venir) » dans l'en-tête
    et leurs cellules sont vides. L'extracteur :
      - lit l'année de référence depuis L4 (« 2025 », « 2026 », …) ;
      - repère le dernier mois publié (n non None) ;
      - calcule la variation Jan → dernier mois publié ;
      - expose `mois_disponibles`, `mois_dernier`, `annee_reference`.

    La variation n'est donc plus systématiquement « Jan → Déc » mais
    « Jan → dernier mois disponible » de l'année courante. Pour la baseline
    annuelle figée (2025), utiliser plutôt `extract_emplois_eerh_annuel`.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Année de référence : en L3 ou L4 selon le millésime du fichier ISQ
    annee_ref = None
    for r in range(2, 6):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        try:
            n = int(str(v).strip())
            if 2000 < n < 2100:
                annee_ref = n
                break
        except (ValueError, TypeError):
            continue

    MOIS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
            'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    month_cols = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25]  # 1-indexed

    rows_iter = list(ws.iter_rows(values_only=True))
    out = []
    i = 7  # data rows start at row 8 (0-indexed = 7)
    while i < len(rows_iter):
        r = rows_iter[i]
        col0, col1 = r[0], r[1]
        if (col0 and isinstance(col0, str) and col0.strip()
                and col1 == 'n'):
            industry_raw = col0
            niveau = _hierarchy_level(industry_raw)
            terminee = _is_terminated(industry_raw)

            scian_match = re.search(r'\(([\d,\s]+)\)\s*(?:\(Terminé\))?\s*$',
                                    industry_raw.replace('\xa0', '').strip())
            scian = scian_match.group(1).replace(' ', '') if scian_match else ''
            label_clean = re.sub(r'\s*\([\d,\s]+\)\s*(?:\(Terminé\))?\s*$', '',
                                 industry_raw.replace('\xa0', '').strip()).strip()

            # n values
            n_series = [_to_num(r[c - 1]) for c in month_cols]
            # TCM row (next)
            tcm_row = rows_iter[i + 1] if i + 1 < len(rows_iter) else (None,) * 26
            tcm_series = [_to_num(tcm_row[c - 1]) for c in month_cols]

            n_jan = n_series[0]
            # Repérer le dernier mois publié (variation Jan → dernier mois)
            n_dernier = None
            mois_dernier_idx = None
            for idx in range(len(n_series) - 1, -1, -1):
                if n_series[idx] is not None:
                    n_dernier = n_series[idx]
                    mois_dernier_idx = idx
                    break
            mois_disponibles = sum(1 for v in n_series if v is not None)

            if n_jan is not None and n_dernier is not None and n_jan != 0:
                variation_pct = (n_dernier - n_jan) / n_jan * 100
            else:
                variation_pct = None

            out.append({
                'scian': scian,
                'industrie': label_clean,
                'niveau': niveau,
                'serie_terminee': terminee,
                'mois': MOIS,
                'n_serie': n_series,
                'tcm_serie': tcm_series,
                'n_janvier': n_jan,
                'n_decembre': n_series[-1],
                'n_dernier_mois': n_dernier,
                'mois_dernier': MOIS[mois_dernier_idx] if mois_dernier_idx is not None else None,
                'mois_disponibles': mois_disponibles,
                'annee_reference': annee_ref,
                'variation_pct': variation_pct,
            })
            i += 2
        else:
            i += 1

    return out


def extract_emplois_eerh_annuel(path: Path) -> list[dict]:
    """
    Tableau EERH série annuelle 2001- (Québec), source des baselines figées.

    Layout : ligne d'années en L5 (col 3 = 2001, col 5 = 2002, ..., paires).
    Pour chaque industrie : deux lignes (n + TCA), avec SCIAN entre
    parenthèses dans le libellé col A et indentation NBSP par sextuplets
    (convention ISQ).

    Retourne une liste de dicts compatible avec extract_emplois_eerh,
    enrichie de `annees`, `n_serie` et `tca_serie`. Les helpers `n_2024`,
    `n_2025` et `tca_2025` exposent directement les chiffres pivots
    utiles à la baseline 2025 du protocole.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Repérer la ligne d'années (2001 en colonne 3)
    annee_row = None
    for r in range(1, 12):
        v = ws.cell(row=r, column=3).value
        if v is None:
            continue
        try:
            if int(str(v).strip()) == 2001:
                annee_row = r
                break
        except (ValueError, TypeError):
            continue
    if annee_row is None:
        raise ValueError("Ligne d'années (2001) introuvable dans le fichier annuel")

    # Lire toutes les années sur colonnes paires
    annees, year_cols = [], []
    c = 3
    while c <= (ws.max_column or 3):
        v = ws.cell(row=annee_row, column=c).value
        if v is not None and str(v).strip():
            try:
                annees.append(int(str(v).strip()))
                year_cols.append(c)
            except (ValueError, TypeError):
                pass
        c += 2

    out = []
    i = annee_row + 1
    while i <= (ws.max_row or annee_row + 1):
        col1 = ws.cell(row=i, column=1).value
        col2 = ws.cell(row=i, column=2).value

        # Ligne d'industrie : libellé + col2 == 'n'
        if (col1 is not None and str(col1).strip()
                and isinstance(col2, str) and col2.strip() == 'n'):
            industry_raw = str(col1)
            niveau = _hierarchy_level(industry_raw)
            terminee = _is_terminated(industry_raw)

            scian_match = re.search(r'\(([\d,\s]+)\)\s*(?:\(Terminé\))?\s*$',
                                    industry_raw.replace('\xa0', '').strip())
            scian = scian_match.group(1).replace(' ', '') if scian_match else ''
            label_clean = re.sub(r'\s*\([\d,\s]+\)\s*(?:\(Terminé\))?\s*$', '',
                                 industry_raw.replace('\xa0', '').strip()).strip()

            n_serie = [_to_num(ws.cell(row=i, column=col).value)
                       for col in year_cols]
            tca_serie = [_to_num(ws.cell(row=i + 1, column=col).value)
                         for col in year_cols] if i + 1 <= (ws.max_row or 0) \
                        else [None] * len(year_cols)

            # Helpers pour les pivots de la baseline 2025
            def _val(an):
                return n_serie[annees.index(an)] if an in annees else None

            def _tca(an):
                return tca_serie[annees.index(an)] if an in annees else None

            out.append({
                'scian': scian,
                'industrie': label_clean,
                'niveau': niveau,
                'serie_terminee': terminee,
                'annees': annees,
                'n_serie': n_serie,
                'tca_serie': tca_serie,
                'n_2024': _val(2024),
                'n_2025': _val(2025),
                'tca_2025': _tca(2025),
            })
            i += 2
        else:
            i += 1

    return out


def extract_ventes_livres(path: Path) -> dict:
    """
    Tableau « Variations mensuelles et annuelles des ventes de livres neufs »
    Layout : A=libellé (hiérarchique avec \\xa0), B=mois courant ($), D=var mois-1 %,
             F=var an-1 mois %, H=cumul YTD ($), J=var cumul an-1 %.
    Le mois et l'année du dépôt sont en R3 et R4.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    annee = ws.cell(row=3, column=1).value or ''
    mois = ws.cell(row=4, column=1).value or ''
    periode = f"{mois} {annee}".strip()

    lignes = []
    for row in ws.iter_rows(min_row=8, max_row=40, values_only=True):
        label_raw = row[0]
        if not label_raw or not str(label_raw).strip():
            continue
        niveau = _hierarchy_level(label_raw)
        label = _clean_label(label_raw)
        # Sauter les lignes purement structurelles (sans valeur)
        if all(_to_num(row[c]) is None for c in (1, 3, 5, 7, 9)):
            # Ligne de section (parent sans données propres) — la garder comme groupe
            lignes.append({
                'libelle': label,
                'niveau': niveau,
                'is_groupe': True,
                'mois_courant': None,
                'var_mois_prec_pct': None,
                'var_an_prec_mois_pct': None,
                'cumul_ytd': None,
                'var_cumul_an_prec_pct': None,
            })
            continue
        lignes.append({
            'libelle': label,
            'niveau': niveau,
            'is_groupe': False,
            'mois_courant': _to_num(row[1]),
            'var_mois_prec_pct': _to_num(row[3]),
            'var_an_prec_mois_pct': _to_num(row[5]),
            'cumul_ytd': _to_num(row[7]),
            'var_cumul_an_prec_pct': _to_num(row[9]),
        })
    return {'periode': periode, 'unite_monetaire': '$ courants', 'lignes': lignes}


def extract_ventes_categorie(path: Path) -> dict:
    """
    Tableau « Ventes de livres neufs selon la catégorie de points de vente,
    données mensuelles » (ISQ tableau 2341).
    Layout : A=libellé hiérarchique ; colonnes B.. = mois ($), dernière = Cumulatif.
    R3 = année, R4 = portée géographique, R6 = en-têtes de colonnes.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']
    annee = ws.cell(row=3, column=1).value or ''
    portee = ws.cell(row=4, column=1).value or ''

    # En-têtes de colonnes (R6) : Janvier, Février, ..., Cumulatif
    headers = []
    for c in range(2, (ws.max_column or 2) + 1):
        v = ws.cell(row=6, column=c).value
        if v is None or str(v).strip() == '':
            break
        headers.append(str(v).strip())
    has_cumul = bool(headers) and headers[-1].lower().startswith('cumul')
    mois = headers[:-1] if has_cumul else headers
    n_mois = len(mois)

    lignes = []
    for row in ws.iter_rows(min_row=8, max_row=60, values_only=True):
        label_raw = row[0]
        if label_raw is None or not str(label_raw).strip():
            if lignes:            # ligne vide après les données -> fin du tableau
                break
            continue
        if str(label_raw).startswith('\n') or str(label_raw).lstrip().startswith('Notes'):
            break                 # bloc de notes -> fin du tableau
        valeurs = [_to_num(row[1 + i]) for i in range(n_mois)]
        cumul = _to_num(row[1 + n_mois]) if has_cumul else None
        lignes.append({
            'libelle': _clean_label(label_raw),
            'niveau': _hierarchy_level(label_raw),
            'is_groupe': all(v is None for v in valeurs) and cumul is None,
            'valeurs': valeurs,
            'cumul': cumul,
        })

    periode = f"Janvier à {mois[-1]} {annee}".strip() if mois else str(annee).strip()
    return {
        'periode': periode,
        'annee': annee,
        'portee': str(portee).strip(),
        'mois': mois,
        'unite_monetaire': '$ courants',
        'lignes': lignes,
    }


def extract_etablissements(path: Path) -> dict:
    """
    Tableau « Nombre d'établissements culturels de certains types »
    Layout : A=libellé hiérarchique, B/D/F/... = années (2004-2024).
    Header année en R4 sous la forme '2004 (ou 2004-2005)'.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Trouver les colonnes-années (R4)
    annees, year_cols = [], []
    for c in ws[4]:
        v = c.value
        if v and re.match(r'\s*\d{4}', str(v)):
            try:
                yr = int(str(v).strip().split()[0])
                annees.append(yr)
                year_cols.append(c.column)
            except (ValueError, IndexError):
                pass

    indicateurs = []
    for row in ws.iter_rows(min_row=6, max_row=42, values_only=True):
        label_raw = row[0]
        if not label_raw or not str(label_raw).strip():
            continue
        niveau = _hierarchy_level(label_raw)
        label = _clean_label(label_raw)
        serie = []
        for yr, col in zip(annees, year_cols):
            val = _to_num(row[col - 1]) if (col - 1) < len(row) else None
            serie.append({'annee': yr, 'valeur': val})
        indicateurs.append({
            'libelle': label,
            'niveau': niveau,
            'serie': serie,
        })
    return {'annees': annees, 'indicateurs': indicateurs}


def extract_indicateurs_cinema(path: Path) -> dict:
    """
    Tableau « Indicateurs des résultats d'exploitation cinémas, données annuelles » (depuis 1975)
    Layout : A=indicateur, B=unité, C/E/G/... = années à partir de 1975.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # Header R4 : C=1975, E=1976, ...
    annees, year_cols = [], []
    for c in ws[4]:
        v = c.value
        if v and isinstance(v, (int, float)) and 1900 < v < 2100:
            annees.append(int(v))
            year_cols.append(c.column)
        elif v and re.match(r'^\d{4}$', str(v).strip()):
            annees.append(int(str(v).strip()))
            year_cols.append(c.column)

    indicateurs = []
    for row in ws.iter_rows(min_row=6, max_row=42, values_only=True):
        label_raw = row[0]
        if not label_raw or not str(label_raw).strip():
            continue
        label = _clean_label(label_raw)
        unite = str(row[1]).strip() if row[1] else ''
        serie = []
        for yr, col in zip(annees, year_cols):
            val = _to_num(row[col - 1]) if (col - 1) < len(row) else None
            serie.append({'annee': yr, 'valeur': val})
        # Filtrer les lignes purement vides
        if all(p['valeur'] is None for p in serie):
            continue
        indicateurs.append({
            'libelle': label,
            'unite': unite,
            'serie': serie,
        })
    return {'annees': annees, 'indicateurs': indicateurs}


# ---------- Helpers spécifiques aux séries cinéma 1985- ----------
#
# Ces trois fichiers (langue, classement, pays annuel) partagent une mise en
# page commune : ligne d'années en colonne paire (B, D, F, ...), libellés en
# colonne A, indentation par NBSP triples (3 NBSP = sous-niveau). C'est une
# convention distincte du tableau « Indicateurs cinéma » (qui démarre en 1975
# et colle ses années en C, E, G, ...).

def _find_year_row_paired(ws, max_scan_row: int = 12) -> int | None:
    """Retourne le numéro de ligne où la colonne B contient une année (1900-2100).
    Pour les tableaux ISQ dont les années sont en colonnes paires B/D/F/...
    """
    for r in range(1, max_scan_row + 1):
        v = ws.cell(row=r, column=2).value
        if v is None:
            continue
        try:
            n = int(str(v).rstrip('¹²³⁴⁵⁶⁷⁸⁹⁰').strip())
            if 1900 < n < 2100:
                return r
        except (ValueError, TypeError):
            continue
    return None


def _read_paired_years(ws, annee_row: int) -> tuple[list[int], list[int]]:
    """Lit la ligne d'années sur colonnes paires (B, D, F, ...).
    Retourne (annees, colonnes_paires) — exclut les colonnes vides en tête/queue.
    """
    annees, cols = [], []
    c = 2
    while c <= (ws.max_column or 2):
        v = ws.cell(row=annee_row, column=c).value
        if v is not None and str(v).strip():
            try:
                annees.append(int(str(v).rstrip('¹²³⁴⁵⁶⁷⁸⁹⁰').strip()))
                cols.append(c)
            except (ValueError, TypeError):
                pass
        c += 2
    return annees, cols


def _indent_level_nbsp3(s) -> int:
    """Niveau hiérarchique pour les fichiers cinéma 1985- (3 NBSP par niveau)."""
    if s is None:
        return 0
    raw = str(s)
    nbsp = len(raw) - len(raw.lstrip('\xa0'))
    return nbsp // 3


def _extract_serie_longue_cinema(path: Path, *,
                                  labels_stop_prefixes: tuple = ('Notes', 'Source', '.. :', '- :'),
                                  ) -> dict:
    """Extracteur commun aux trois nouveaux tableaux cinéma 1985-.

    Logique : trouver la ligne d'années, puis lire chaque ligne suivante comme
    un indicateur (libellé col A, valeurs aux colonnes paires). On s'arrête
    quand on rencontre une ligne de notes/sources ou une ligne vide après
    avoir commencé à accumuler des indicateurs.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    annee_row = _find_year_row_paired(ws)
    if annee_row is None:
        raise ValueError(f"Ligne d'années introuvable dans {path.name}")
    annees, year_cols = _read_paired_years(ws, annee_row)

    indicateurs = []
    for r in range(annee_row + 1, (ws.max_row or annee_row + 1) + 1):
        label_raw = ws.cell(row=r, column=1).value
        if label_raw is None or not str(label_raw).strip():
            # Ligne vide : si on a déjà accumulé des indicateurs, c'est la fin
            if indicateurs:
                break
            continue
        label_str = str(label_raw).lstrip(' \xa0').rstrip()
        if any(label_str.startswith(p) for p in labels_stop_prefixes):
            break
        niveau = _indent_level_nbsp3(label_raw)
        serie = [{'annee': yr, 'valeur': _to_num(ws.cell(row=r, column=col).value)}
                 for yr, col in zip(annees, year_cols)]
        # Filtrer les lignes sans aucune valeur (sépareurs visuels)
        if all(p['valeur'] is None for p in serie):
            continue
        indicateurs.append({
            'libelle': _clean_label(label_raw),
            'niveau': niveau,
            'serie': serie,
        })
    return {'annees': annees, 'indicateurs': indicateurs}


# ---------- Nouveaux extracteurs cinéma (publiés juin 2026) ----------

def extract_cinema_langue(path: Path) -> dict:
    """Tableau « Résultats d'exploitation cinéma — langue de projection » (1985-).

    Hiérarchie attendue :
      Projections
      Langue française
        Cinémas
        Ciné-parcs
      Autres que le français
        Cinémas
        Ciné-parcs

    Indicateur direct du marché francophone — pertinent pour la Loi 109.
    """
    return _extract_serie_longue_cinema(path)


def extract_cinema_classement(path: Path) -> dict:
    """Tableau « Résultats d'exploitation cinéma — catégorie de classement » (1985-).

    Hiérarchie attendue : Projections / Visa général / 13 ans et plus /
    16 ans et plus / 18 ans et plus, chacune subdivisée Cinémas / Ciné-parcs
    (sauf 18+).
    """
    return _extract_serie_longue_cinema(path)


def extract_cinema_pays_annuel(path: Path) -> dict:
    """Tableau « Résultats d'exploitation cinéma — pays d'origine, données annuelles » (1985-).

    Liste plate de pays, indicateur = assistance (spectateurs) ventilée par
    pays d'origine du film. Permet de calculer directement C_cinema dans R3
    sans recourir à une multiplication recettes × part hebdomadaire.

    Retourne en plus du format standard une clé `assistance_par_pays` qui
    indexe les indicateurs par libellé de pays pour usage rapide en aval.
    """
    base = _extract_serie_longue_cinema(path)
    # Index complémentaire par pays
    base['assistance_par_pays'] = {
        ind['libelle']: ind['serie']
        for ind in base['indicateurs']
    }
    return base


def extract_remunerations_eerh_statcan(path: Path) -> dict:
    """Table StatCan CANSIM 14-10-0223 « Emploi et rémunération hebdomadaire
    moyenne (incluant le temps supplémentaire) pour l'ensemble des salariés
    selon la province et le territoire, données mensuelles, désaisonnalisées ».

    Format : zip contenant un CSV en format long (≈ 47 Mo non compressé) avec
    les colonnes PÉRIODE DE RÉFÉRENCE, GÉO, Estimation, SCIAN, ..., VALEUR.
    L'extracteur stream-parse le CSV en filtrant uniquement les lignes Québec
    × secteurs SCIAN suivis (niveau 2 chiffres).

    **Limite de granularité importante** : la table n'est diffusée qu'au niveau
    SCIAN 2 chiffres pour les coupes provinciales. Les secteurs culturels
    précis qu'on suit ailleurs dans le pipeline (5121 film, 5162 streaming,
    7111 arts d'interprétation, etc.) sont ici agrégés dans [51] et [71].
    C'est documenté dans le champ `note_granularite` du retour.

    Sert la lentille 3 « rémunération » de l'analyse AI-exposure, en
    complément des effectifs de `emplois_eerh_annuel` (ISQ, niveau 4 chiffres).
    """
    # Secteurs SCIAN à 2 chiffres pertinents pour la lentille 3
    SECTEURS = {
        '51': "Industrie de l'information et industrie culturelle",
        '71': 'Arts, spectacles et loisirs',
    }
    # Libellés StatCan des deux estimations (matching par sous-chaîne robuste)
    EST_EMPLOI = 'Emploi'
    EST_REMU = 'Rémunération'

    # Collecte : (code, mesure_key) -> liste [(periode, valeur)]
    serie = defaultdict(list)

    with zipfile.ZipFile(path) as z:
        csv_name = [n for n in z.namelist()
                    if n.endswith('.csv') and 'MetaData' not in n]
        if not csv_name:
            raise ValueError(f"Aucun CSV de données dans {path.name}")
        with z.open(csv_name[0]) as fh:
            text = io.TextIOWrapper(fh, encoding='utf-8-sig', newline='')
            reader = csv.reader(text, delimiter=';')
            next(reader)  # header
            for row in reader:
                if len(row) < 12:
                    continue
                periode, geo, _, estim, scian = row[:5]
                if geo != 'Québec':
                    continue
                m = re.search(r'\[([^\]]+)\]', scian)
                code = m.group(1).strip() if m else ''
                if code not in SECTEURS:
                    continue
                if EST_EMPLOI in estim:
                    mesure = 'effectifs'
                elif EST_REMU in estim:
                    mesure = 'remuneration_hebdo'
                else:
                    continue
                val = _to_num(row[11])
                serie[(code, mesure)].append((periode, val))

    # Agrégations annuelles (moyenne arithmétique mensuelle)
    moyennes = defaultdict(dict)  # (code, mesure) -> {annee: moyenne}
    n_mois = defaultdict(dict)    # (code, mesure) -> {annee: nombre_mois}
    for key, observations in serie.items():
        by_year = defaultdict(list)
        for per, v in observations:
            if v is None:
                continue
            try:
                an = int(per.split('-')[0])
                by_year[an].append(v)
            except (ValueError, IndexError):
                pass
        for an, vals in by_year.items():
            moyennes[key][an] = round(sum(vals) / len(vals), 2)
            n_mois[key][an] = len(vals)

    # Détecter la plage globale
    toutes_periodes = sorted({p for obs in serie.values() for p, _ in obs})
    periode_min = toutes_periodes[0] if toutes_periodes else None
    periode_max = toutes_periodes[-1] if toutes_periodes else None

    # Assembler la sortie
    secteurs_out = []
    for code, libelle in SECTEURS.items():
        bloc = {'code_scian': code, 'libelle': libelle, 'mesures': {}}
        for mesure in ('effectifs', 'remuneration_hebdo'):
            key = (code, mesure)
            if key not in serie:
                continue
            bloc['mesures'][mesure] = {
                'serie_mensuelle': [
                    {'periode': p, 'valeur': v} for p, v in serie[key]
                ],
                'moyennes_annuelles': [
                    {'annee': an, 'valeur': moyennes[key][an],
                     'n_mois': n_mois[key][an]}
                    for an in sorted(moyennes[key].keys())
                ],
            }
        secteurs_out.append(bloc)

    return {
        'source': ('Statistique Canada — CANSIM 14-10-0223 (Emploi et rémunération '
                   'hebdomadaire moyenne, EERH, données mensuelles désaisonnalisées)'),
        'tableau': '14-10-0223',
        'periode_min': periode_min,
        'periode_max': periode_max,
        'note_granularite': ('SCIAN niveau 2 chiffres uniquement pour les coupes '
                             'provinciales. Les sous-secteurs culture précis '
                             '(5121 film, 5162 streaming, 7111 arts d\'interprétation, etc.) '
                             'sont agrégés ici dans [51] et [71]. Pour la granularité '
                             'fine, utiliser emplois_eerh_annuel (ISQ).'),
        'secteurs': secteurs_out,
    }


def extract_ai_exposure_culture(path: Path) -> dict:
    """Indice C-AIOE (complementarity-adjusted AI occupational exposure) pour
    les industries culturelles canadiennes, d'après Mehdi, Allen, Lesica & Watt
    (Statistique Canada, mars 2026).

    Sert la **sous-lentille 1a « demande experte »** de l'analyse AI-exposure.
    Mesure prospective : pour chaque industrie culturelle, pourcentage des
    emplois classés en trois catégories selon l'indice de Felten et al. (2021)
    et Pizzinelli et al. (2023) :

      * **HE_LC** = Haute exposition + Faible complémentarité → potentiel de
        substitution AI (l'IA peut remplacer le travail humain)
      * **HE_HC** = Haute exposition + Haute complémentarité → potentiel
        d'augmentation (l'IA enrichit le travail humain)
      * **LE**    = Faible exposition à l'IA (peu de transformation prévue)

    Données 2021 (recensement + CEEDD), classification NAICS 2022, sexe (men+
    et women+ depuis 2021). Source : Statistique Canada, Economic and Social
    Reports, 25 mars 2026, sous licence ouverte.

    Granularité : Canada national. Pas de coupe Québec disponible dans la
    publication ; limite cohérente avec la lentille 2 AEI (Anthropic Economic
    Index Canada).
    """
    industries_data = defaultdict(lambda: {'men+': None, 'women+': None})

    with open(path, encoding='utf-8') as fh:
        reader = csv.reader(fh)
        for row in reader:
            if not row or row[0].startswith('#'):
                continue
            if row[0] == 'industrie_code':  # header
                continue
            if len(row) < 6:
                continue
            code, libelle, sexe, he_lc, he_hc, le = row[:6]
            if not code:
                continue
            scores = {
                'he_lc_pct': float(he_lc) if he_lc.strip() else None,
                'he_hc_pct': float(he_hc) if he_hc.strip() else None,
                'le_pct': float(le) if le.strip() else None,
            }
            # libellé est répété ; on garde
            industries_data[(code, libelle)][sexe] = scores

    # Reformater en liste d'industries
    industries_out = []
    for (code, libelle), par_sexe in industries_data.items():
        # Calcul d'agrégat hommes+femmes (moyenne simple, faute d'avoir les
        # parts de chaque sexe dans l'industrie) — à interpréter avec prudence
        h = par_sexe.get('men+') or {}
        f = par_sexe.get('women+') or {}
        def _moy(k):
            vh, vf = h.get(k), f.get(k)
            if vh is None and vf is None:
                return None
            if vh is None:
                return vf
            if vf is None:
                return vh
            return round((vh + vf) / 2, 1)
        industries_out.append({
            'code': code,
            'libelle': libelle,
            'men+': par_sexe.get('men+'),
            'women+': par_sexe.get('women+'),
            'moyenne_sexes': {
                'he_lc_pct': _moy('he_lc_pct'),
                'he_hc_pct': _moy('he_hc_pct'),
                'le_pct': _moy('le_pct'),
            },
        })

    # Tri : industries culturelles d'abord (par taux HE_LC moyen décroissant),
    # puis « autres industries » en référence à la fin
    def _tri_key(ind):
        is_autre = ind['code'] == 'autres'
        he_lc = ind['moyenne_sexes']['he_lc_pct'] or 0
        return (is_autre, -he_lc)
    industries_out.sort(key=_tri_key)

    return {
        'source': ('Statistique Canada — Mehdi, Allen, Lesica & Watt (2026), '
                   '« Potential occupational exposure to artificial intelligence '
                   'across selected cultural industries in Canada », Economic and '
                   'Social Reports, 25 mars 2026'),
        'doi': '10.25318/36280001202600300003-eng',
        'url': ('https://www150.statcan.gc.ca/n1/pub/36-28-0001/2026003/article/'
                '00003-eng.htm'),
        'methode': ('Indice C-AIOE (complementarity-adjusted AI occupational '
                    'exposure), Felten, Raj & Seamans (2021) et Pizzinelli et '
                    'al. (2023). Classification des emplois en trois catégories : '
                    'HE_LC (haute exposition + faible complémentarité → '
                    'substitution potentielle), HE_HC (haute exposition + haute '
                    'complémentarité → augmentation potentielle), LE (faible '
                    'exposition à l\'IA).'),
        'periode_reference': 'Mai 2021 (Census + CEEDD)',
        'pays': 'Canada (national)',
        'note_limites': ('Granularité Canada national, pas de coupe Québec dans '
                         "la publication. Sample : employés 18-64 ans, secteur "
                         "commercial seulement (hors admin. publique, services "
                         "éducatifs, santé). Estimation basée sur la faisabilité "
                         "technologique de remplacer les tâches, pas sur le "
                         "comportement réel des employeurs. Exposition n'implique "
                         "pas perte d'emploi : transformation potentielle des "
                         "tâches uniquement. Données 2021, à actualiser quand "
                         "StatCan publiera la prochaine vague."),
        'industries': industries_out,
    }


def extract_job_vacancy_quebec(path: Path) -> dict:
    """Table StatCan CANSIM 14-10-0442 — Postes vacants, employés salariés,
    taux de postes vacants et salaire horaire offert moyen, par SCIAN 3 chiffres,
    province ou territoire, données trimestrielles non désaisonnalisées.

    Sert la **sous-lentille 1b « demande marché »** de l'analyse AI-exposure :
    ce que les employeurs québécois cherchent réellement à embaucher dans les
    industries culturelles, à quel salaire, à quel taux de vacance.

    Filtre : GÉO == 'Québec' + SCIAN dans {512, 513, 515, 516, 519, 711, 712}.
    Format zip / CSV long, identique au pattern de CANSIM 14-10-0223.

    Métriques exposées par SCIAN :
      - postes_vacants : nombre absolu de postes ouverts (Statistique « Postes vacants »)
      - employes_salaries : nombre d'emplois salariés en place
      - taux_postes_vacants : ratio postes vacants / (postes vacants + employés)
      - salaire_horaire_offert : moyenne du salaire horaire offert dans les postes
        vacants, en $ courants
    """
    SECTEURS = {
        '512': 'Industries du film et de l\'enregistrement sonore',
        '513': 'Édition',
        '515': 'Radiotélévision (sauf par Internet)',
        '516': 'Radiotélévision et fournisseurs de contenu par Internet',
        '519': 'Portails de recherche Web, bibliothèques, archives',
        '711': 'Arts d\'interprétation, sports-spectacles et activités connexes',
        '712': 'Établissements du patrimoine',
    }
    STAT_MAPPING = {
        'Postes vacants': 'postes_vacants',
        'Employés salariés': 'employes_salaries',
        'Taux de postes vacants': 'taux_postes_vacants',
        'Moyenne du salaire horaire offert': 'salaire_horaire_offert',
    }

    # (code, periode) -> {stat_key: value}
    obs = defaultdict(lambda: defaultdict(lambda: None))

    with zipfile.ZipFile(path) as z:
        csv_name = [n for n in z.namelist()
                    if n.endswith('.csv') and 'MetaData' not in n]
        if not csv_name:
            raise ValueError(f"Aucun CSV de données dans {path.name}")
        with z.open(csv_name[0]) as fh:
            text = io.TextIOWrapper(fh, encoding='utf-8-sig', newline='')
            reader = csv.reader(text, delimiter=';')
            next(reader)  # header
            for row in reader:
                if len(row) < 12:
                    continue
                periode, geo, _, scian, stat = row[:5]
                if geo != 'Québec':
                    continue
                m = re.search(r'\[([^\]]+)\]', scian)
                code = m.group(1).strip() if m else ''
                if code not in SECTEURS:
                    continue
                stat_key = STAT_MAPPING.get(stat)
                if stat_key is None:
                    continue
                val = _to_num(row[11])
                obs[(code, periode)][stat_key] = val

    # Plage de périodes disponibles
    toutes_periodes = sorted({p for (_, p) in obs.keys()})
    periode_min = toutes_periodes[0] if toutes_periodes else None
    periode_max = toutes_periodes[-1] if toutes_periodes else None
    # Les 5 derniers trimestres pour les moyennes
    derniers_5 = toutes_periodes[-5:] if len(toutes_periodes) >= 5 else toutes_periodes

    # Assembler par secteur
    secteurs_out = []
    for code, libelle in SECTEURS.items():
        periodes_secteur = sorted({p for (c, p) in obs.keys() if c == code})
        if not periodes_secteur:
            # Secteur non couvert (515 archivé par exemple)
            secteurs_out.append({
                'code_scian': code,
                'libelle': libelle,
                'statut': 'non_couvert',
                'note': f"Aucune observation disponible pour SCIAN [{code}] au Québec dans la table.",
            })
            continue

        serie = [
            {
                'periode': p,
                **{k: obs[(code, p)].get(k) for k in STAT_MAPPING.values()},
            }
            for p in periodes_secteur
        ]

        # Moyennes sur les 5 derniers trimestres
        def _moy(stat_key):
            vals = [obs[(code, p)].get(stat_key) for p in derniers_5
                    if obs[(code, p)].get(stat_key) is not None]
            return round(sum(vals) / len(vals), 2) if vals else None

        moyennes = {
            'postes_vacants': _moy('postes_vacants'),
            'employes_salaries': _moy('employes_salaries'),
            'taux_postes_vacants': _moy('taux_postes_vacants'),
            'salaire_horaire_offert': _moy('salaire_horaire_offert'),
            'n_trimestres': len(derniers_5),
            'periodes_couvertes': derniers_5,
        }

        secteurs_out.append({
            'code_scian': code,
            'libelle': libelle,
            'serie_trimestrielle': serie,
            'moyennes_5_derniers_trimestres': moyennes,
        })

    return {
        'source': ('Statistique Canada — CANSIM 14-10-0442 (Postes vacants, '
                   'employés salariés, taux de postes vacants et salaire horaire '
                   'offert moyen selon le sous-secteur SCIAN 3 chiffres, '
                   'données trimestrielles non désaisonnalisées)'),
        'tableau': '14-10-0442-01',
        'periode_min': periode_min,
        'periode_max': periode_max,
        'note_limites': ('Granularité SCIAN 3 chiffres (sous-secteur), entre les '
                         'effectifs ISQ niveau 4 et la rémunération EERH niveau 2. '
                         "SCIAN 515 (Radiotélévision sauf par Internet) archivé "
                         "côté ISQ comme côté StatCan, série non disponible. "
                         'Données non désaisonnalisées — interpréter les '
                         'variations trimestrielles avec prudence. La métrique '
                         '« salaire horaire offert » correspond au salaire que '
                         "l'employeur affiche dans le poste vacant, pas au "
                         'salaire effectivement versé aux salariés en place.'),
        'secteurs': secteurs_out,
    }


def extract_aei_canada(path: Path) -> dict:
    """Anthropic Economic Index — données Claude.ai pour le Canada.

    Source : `aei_raw_claude_ai_<period>.csv` (release Hugging Face Anthropic/EconomicIndex).
    Format : CSV en format long, ~478k lignes. L'extracteur filtre uniquement
    les observations Canada (`geo_id == 'CA'`).

    Sert la **lentille 2 « usage révélé »** de l'analyse AI-exposure : ce que
    les utilisateurs canadiens de Claude.ai font réellement, ventilation par
    tâche O*NET et par mode de collaboration (directive, learning, feedback
    loop, etc.).

    Limite assumée : l'intersection `onet_task::collaboration` n'est exposée
    par Anthropic que pour la géographie GLOBAL (filtre vie privée). Pour le
    Canada, on a donc le breakdown collaboration *agrégé toutes tâches
    confondues*, sans ventilation par tâche.

    Périmètre créatif au sens du Carnet (filtrage par sous-chaînes signatures
    des libellés O*NET) :
      * **Cœur culturel** (7 tâches) : illustration, animation/VFX, art visuel,
        design mode, édition de manuscrits, critique d'œuvres
      * **Création de contenu écrit** (6 tâches) : rédaction publicitaire,
        édition de copie, design éditorial, web content
    """
    # Sous-chaînes signatures des libellés O*NET retenus comme créatifs
    COEUR_CULTUREL_PATTERNS = [
        'write reviews of literary',
        'read, evaluate and edit manuscripts',
        'program computerized graphic effects',
        'develop, copy, or adapt designs for garments',
        'create finished art work',
        'design complex graphics and animation',
        'create custom illustrations',
    ]
    CONTENU_ECRIT_PATTERNS = [
        'prepare, rewrite and edit copy',
        'edit, standardize, or make changes to material prepared by other writers',
        'write advertising copy',
        'edit or rewrite existing copy',
        'develop briefings, brochures, multimedia presentations',
        'write, design, or edit web page content',
    ]

    def _classer_tache(libelle: str) -> str | None:
        low = libelle.lower()
        for p in COEUR_CULTUREL_PATTERNS:
            if p in low:
                return 'coeur_culturel'
        for p in CONTENU_ECRIT_PATTERNS:
            if p in low:
                return 'contenu_ecrit'
        return None

    # Collecte
    taches_ca = {}  # libelle -> {'pct': ..., 'count': ...}
    collaboration_ca = {}  # cluster -> {'pct': ..., 'count': ...}
    date_start = None
    date_end = None
    platform = None

    with open(path, encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if row['geo_id'] != 'CA':
                continue
            date_start = date_start or row['date_start']
            date_end = date_end or row['date_end']
            platform = platform or row['platform_and_product']

            if row['facet'] == 'onet_task':
                t = row['cluster_name']
                if t not in taches_ca:
                    taches_ca[t] = {'pct': None, 'count': None}
                val = float(row['value']) if row['value'] else None
                if row['variable'] == 'onet_task_pct':
                    taches_ca[t]['pct'] = val
                elif row['variable'] == 'onet_task_count':
                    taches_ca[t]['count'] = val
            elif row['facet'] == 'collaboration':
                c = row['cluster_name']
                if c not in collaboration_ca:
                    collaboration_ca[c] = {'pct': None, 'count': None}
                val = float(row['value']) if row['value'] else None
                if row['variable'] == 'collaboration_pct':
                    collaboration_ca[c]['pct'] = val
                elif row['variable'] == 'collaboration_count':
                    collaboration_ca[c]['count'] = val

    # Agrégats collaboration (productif vs apprentissage selon la grille AI-exposure)
    # Productif = directive + feedback loop + task iteration (substitution potentielle)
    # Apprentissage = learning + validation (augmentation du capital humain)
    PRODUCTIF_MODES = {'directive', 'feedback loop', 'task iteration'}
    APPRENTISSAGE_MODES = {'learning', 'validation'}
    productif_pct = sum(collaboration_ca.get(m, {}).get('pct') or 0
                        for m in PRODUCTIF_MODES)
    apprentissage_pct = sum(collaboration_ca.get(m, {}).get('pct') or 0
                            for m in APPRENTISSAGE_MODES)
    none_pct = collaboration_ca.get('none', {}).get('pct') or 0
    ratio = round(productif_pct / apprentissage_pct, 2) if apprentissage_pct > 0 else None

    # Classer les tâches en périmètre créatif
    coeur_culturel = []
    contenu_ecrit = []
    for libelle, data in taches_ca.items():
        cat = _classer_tache(libelle)
        if cat is None:
            continue
        item = {
            'tache_onet': libelle,
            'pct_total_canada': data['pct'],
            'count': data['count'],
            'categorie': cat,
        }
        if cat == 'coeur_culturel':
            coeur_culturel.append(item)
        else:
            contenu_ecrit.append(item)

    coeur_culturel.sort(key=lambda x: -(x['pct_total_canada'] or 0))
    contenu_ecrit.sort(key=lambda x: -(x['pct_total_canada'] or 0))

    pct_total_coeur = sum(t['pct_total_canada'] or 0 for t in coeur_culturel)
    pct_total_contenu = sum(t['pct_total_canada'] or 0 for t in contenu_ecrit)

    return {
        'source': ('Anthropic Economic Index — Claude.ai Free/Pro/Max, '
                   'release Hugging Face Anthropic/EconomicIndex, cinquième édition '
                   '(« Learning curves », publiée 2026-03-24)'),
        'release_anthropic': '2026-03-24',
        'periode_start': date_start,
        'periode_end': date_end,
        'platform': platform,
        'pays': 'CA',
        'note_limites': ("Granularité géographique : Canada national, pas de "
                         "coupe Québec disponible dans la release. L'intersection "
                         "onet_task::collaboration n'est exposée par Anthropic que "
                         "pour la géographie GLOBAL (filtre vie privée) ; pour le "
                         "Canada on a donc le breakdown collaboration agrégé toutes "
                         "tâches confondues, pas par tâche. Période d'observation : "
                         "une seule semaine (5-12 février 2026) ; pas de tendance "
                         "temporelle possible avec ce seul fichier."),
        'collaboration_canada': {
            mode: collaboration_ca.get(mode, {}).get('pct')
            for mode in ('directive', 'task iteration', 'learning',
                         'feedback loop', 'validation', 'none', 'not_classified')
        },
        'agregats_collaboration': {
            'productif_pct': round(productif_pct, 2),
            'apprentissage_pct': round(apprentissage_pct, 2),
            'none_pct': round(none_pct, 2),
            'ratio_productif_apprentissage': ratio,
        },
        'taches_creatives': {
            'coeur_culturel': coeur_culturel,
            'contenu_ecrit': contenu_ecrit,
            'pct_total_coeur_culturel': round(pct_total_coeur, 2),
            'pct_total_contenu_ecrit': round(pct_total_contenu, 2),
            'pct_total_creatif': round(pct_total_coeur + pct_total_contenu, 2),
        },
        'meta': {
            'n_taches_onet_canada_total': len(taches_ca),
            'n_taches_coeur_culturel': len(coeur_culturel),
            'n_taches_contenu_ecrit': len(contenu_ecrit),
        },
    }


def extract_ventes_livres_numeriques(path: Path) -> dict:
    """
    Tableau ISQ « Ventes de livres numériques, données annuelles, Québec »
    (URL permanente : statistique.quebec.ca/fr/produit/tableau/3408).
    Méthodologie : Optique culture no 41. Inclut les ventes gratuites et autoédités.
    Valeur exprimée au prix payé par le consommateur avant taxes.

    Layout :
      L1 = titre, L4 = en-tête années (cols pairs 3,5,7... = années ; cols impairs = marqueurs ISQ)
      L6 = Nombre d'exemplaires (unité n)
      L7 = Valeur des ventes (unité $)
      L8 = Prix moyen (unité $)
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['Tableau']

    # En-tête : lire les années dans la ligne 4 (colonnes paires à partir de 3)
    annees = []
    cols_annees = []
    for c in range(3, (ws.max_column or 3) + 1):
        v = ws.cell(row=4, column=c).value
        if v is None or str(v).strip() == '':
            continue
        try:
            an = int(str(v).strip())
            annees.append(an)
            cols_annees.append(c)
        except ValueError:
            continue

    # Trois métriques fixes : exemplaires (L6), valeur (L7), prix moyen (L8)
    def _serie(row_num):
        return [_to_num(ws.cell(row=row_num, column=c).value) for c in cols_annees]

    series = {
        'exemplaires': {
            'libelle': "Nombre d'exemplaires",
            'unite': 'n',
            'valeurs': _serie(6),
        },
        'valeur_ventes': {
            'libelle': 'Valeur des ventes',
            'unite': '$',
            'valeurs': _serie(7),
        },
        'prix_moyen': {
            'libelle': 'Prix moyen',
            'unite': '$',
            'valeurs': _serie(8),
        },
    }

    # Métadonnées éditoriales : lien permanent et date de maj dans les notes
    notes_raw = ws.cell(row=11, column=1).value or ''
    lien = ''
    maj = ''
    for line in str(notes_raw).split('\n'):
        line = line.strip()
        if 'statistique.quebec.ca' in line:
            lien = line
        # capture du format "29 juin 2026" sous "Mise à jour :"
    # Lecture explicite des notes pour maj
    full_notes = str(notes_raw)
    if 'Mise à jour' in full_notes:
        suffix = full_notes.split('Mise à jour', 1)[1]
        for line in suffix.split('\n'):
            line = line.strip(' :\t')
            if line and not line.startswith('Mise'):
                maj = line
                break

    return {
        'titre': 'Ventes de livres numériques, données annuelles, Québec',
        'source': 'Institut de la statistique du Québec (Observatoire de la culture et des communications du Québec)',
        'lien_permanent': lien or 'statistique.quebec.ca/fr/produit/tableau/3408',
        'mise_a_jour': maj,
        'periode': f"{annees[0]}-{annees[-1]}" if annees else '',
        'annees': annees,
        'series': series,
        'note_methodo': "Méthodologie : Optique culture no 41. Valeur au prix payé par le consommateur avant taxes. Inclut les ventes gratuites et les autoédités.",
    }


# ---------- Registry ----------

EXTRACTORS = {
    'extract_part_qc': extract_part_qc,
    'extract_volume_musique': extract_volume_musique,
    'extract_cinema_pays': extract_cinema_pays,
    'extract_palmares': extract_palmares,
    'extract_evolution': extract_evolution,
    'extract_emplois_eerh': extract_emplois_eerh,
    'extract_ventes_livres': extract_ventes_livres,
    'extract_ventes_categorie': extract_ventes_categorie,
    'extract_etablissements': extract_etablissements,
    'extract_indicateurs_cinema': extract_indicateurs_cinema,
    'extract_cinema_langue': extract_cinema_langue,
    'extract_cinema_classement': extract_cinema_classement,
    'extract_cinema_pays_annuel': extract_cinema_pays_annuel,
    'extract_emplois_eerh_annuel': extract_emplois_eerh_annuel,
    'extract_remunerations_eerh_statcan': extract_remunerations_eerh_statcan,
    'extract_aei_canada': extract_aei_canada,
    'extract_job_vacancy_quebec': extract_job_vacancy_quebec,
    'extract_ai_exposure_culture': extract_ai_exposure_culture,
    'extract_ventes_livres_numeriques': extract_ventes_livres_numeriques,
}
